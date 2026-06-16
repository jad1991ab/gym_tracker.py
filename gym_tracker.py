import streamlit as st
import pandas as pd
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# إعدادات الصفحة
st.set_page_config(page_title="مفكرة التمارين", layout="wide", initial_sidebar_state="collapsed")

# تطبيق نمط RTL (من اليمين إلى اليسار) للواجهة العربية
st.markdown("""
    <style>
    body, div, p, h1, h2, h3, h4, h5, h6, label, input, button, select {
        direction: RTL;
        text-align: right;
    }
    .stButton>button {
        width: 100%;
    }
    </style>
    """, unsafe_allow_html=True)

# الاتصال بقاعدة بيانات جوجل شيتس السحابية
# ملاحظة: سيقوم Streamlit بالاتصال تلقائياً عبر الإعدادات السحابية
try:
    conn = st.connection("gsheets", type="sheets")
    # قراءة البيانات (تأكد من تسمية الورقة الأولى بـ Sheet1 أو تعديلها بالأسفل)
    df = conn.read(ttl="0")
    # إذا كانت الورقة فارغة، نقوم بتجهيز الأعمدة الأساسية
    if df.empty or len(df.columns) < 5:
        df = pd.DataFrame(columns=["المعرف", "التاريخ", "نوع التمرين", "المدة (دقائق)", "ملاحظات"])
except Exception as e:
    st.error("جاري الاتصال بقاعدة البيانات السحابية أو إعدادها لأول مرة...")
    df = pd.DataFrame(columns=["المعرف", "التاريخ", "نوع التمرين", "المدة (دقائق)", "ملاحظات"])

st.title("🏋️‍♂️ مفكرة التمارين الرياضية السحابية")
st.write("سجل تمارينك من الموبايل في النادي وتابعها من كمبيوترك في أي وقت!")

# سحب جانبي لإدخال البيانات
with st.sidebar:
    st.header("📝 تسجيل نشاط جديد")
    
    # نموذج الإدخال
    input_date = st.date_input("تاريخ التمرين", datetime.now().date())
    exercise_type = st.selectbox("نوع التمرين", ["حديد / مقاومة", "كارديو (جري/مشي)", "سباحة", "كرة قدم", "لياقة بدنية / سويدي", "أخرى"])
    duration = st.number_input("المدة (بالدقائق)", min_value=1, max_value=300, value=45, step=5)
    notes = st.text_area("ملاحظات إضافية", placeholder="مثال: تمرین رجلين، شدة عالية...")
    
    submit_button = st.button("💾 حفظ التمرين سحابياً")

# معالجة إضافة تمرين جديد
if submit_button:
    # توليد معرف فريد يعتمد على الوقت الحالي لمنع التكرار وتسهيل الحذف
    unique_id = datetime.now().strftime("%Y%m%d%H%M%S")
    
    new_data = pd.DataFrame({
        "المعرف": [unique_id],
        "التاريخ": [str(input_date)],
        "نوع التمرين": [exercise_type],
        "المدة (دقائق)": [int(duration)],
        "ملاحظات": [notes]
    })
    
    # دمج البيانات وتحديث الجوجل شيتس سحابياً
    updated_df = pd.concat([df, new_data], ignore_value=True) if not df.empty else new_data
    
    try:
        conn.update(data=updated_df)
        st.sidebar.success("✅ تم حفظ التمرين في قاعدة البيانات السحابية!")
        st.rerun()
    except Exception as e:
        st.sidebar.error(f"حدث خطأ أثناء الحفظ السحابي: {e}")

# عرض البيانات والإحصائيات في الصفحة الرئيسية
st.header("📊 سجل التمارين المسجلة")

if not df.empty and len(df) > 0:
    # عرض إحصائيات سريعة
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("إجمالي التمارين", f"{len(df)} تمارين")
    with col2:
        st.metric("مجموع الدقائق الرياضية", f"{int(df['المدة (دقائق)'].astype(int).sum())} دقيقة")
    with col3:
        st.metric("أكثر تمرين تكراراً", df["نوع التمرين"].mode()[0] if not df["نوع التمرين"].empty else "غير محدد")
    
    st.write("---")
    
    # عرض الجدول مع زر حذف لكل سطر
    # نقوم بعمل حلقة لعرض الأسطر بشكل منسق مع خيار الحذف
    for index, row in df.iterrows():
        # إنشاء حاوية لكل تمرين تظهر بشكل بطاقة منسقة
        with st.container():
            c1, c2, c3, c4, c5 = st.columns([2, 3, 2, 4, 1.5])
            with c1:
                st.write(f"📅 **{row['التاريخ']}**")
            with c2:
                st.write(f"💪 **{row['نوع التمرين']}**")
            with c3:
                st.write(f"⏱️ **{row['المدة (دقائق)']} دقيقة**")
            with c4:
                st.write(f"📝 {row['ملاحظات'] if pd.notna(row['ملاحظات']) else '-'}")
            with c5:
                # زر الحذف يعتمد على المعرف الفريد للسطر
                delete_clicked = st.button("🗑️ حذف", key=f"del_{row['المعرف']}")
                if delete_clicked:
                    # حذف السطر بناءً على المعرف فريد
                    updated_df = df[df["المعرف"] != row["المعرف"]]
                    try:
                        conn.update(data=updated_df)
                        st.success("🗑️ تم حذف التمرين بنجاح سحابياً!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"خطأ أثناء الحذف: {e}")
        st.markdown("<hr style='margin:0.5em 0px; border-color:#eee;'>", unsafe_allow_html=True)

    # 📥 بناء ملف الإكسل الاحترافي المنسق للتحميل
    wb = Workbook()
    ws = wb.active
    ws.title = "سجل التمارين الرياضية"
    ws.views.sheetView[0].rightToLeft = True  # جعل الملف من اليمين إلى اليسار ناتيف

    # تجهيز الداتا فريم للتصدير بدون عمود المعرف الفريد لإبقاء المظهر نظيفاً
    export_df = df.drop(columns=["المعرف"])
    
    # كتابة العناوين وتنسيقها
    for r in dataframe_to_rows(export_df, index=False, header=True):
        ws.append(r)

    # الألوان والتنسيقات
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=11, color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # تطبيق التنسيق على الصف الأول (العناوين)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    # تطبيق التنسيقات على بقية أسطر البيانات
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            if cell.column in [1, 3]:  # التاريخ والمدة في المنتصف
                cell.alignment = center_align
            else:
                cell.alignment = right_align
        ws.row_dimensions[row[0].row].height = 22

    # ضبط تلقائي لعرض الأعمدة بناءً على المحتوى لتجنب اختفاء النصوص
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    # تحويل ملف الإكسل المنسق إلى سيل من البايتات ليدعم تحميله عبر المتصفح مباشرة
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    st.write(" ")
    st.download_button(
        label="📥 تحميل سجل التمارين كاملاً كملف Excel منسق",
        data=excel_buffer,
        file_name=f"gym_activities_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("💡 لا توجد تمارين مسجلة سحابياً حتى الآن. ابدأ بتسجيل أول تمرين لك من القائمة الجانبية!")
